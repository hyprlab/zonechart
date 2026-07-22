"""Parse a UPS zone chart workbook (.xls/.xlsx) into a JSON-ready dict.

UPS publishes one chart per origin 3-digit ZIP prefix. The sheet is a table of
destination 3-digit prefixes with one zone code per service, followed by
footnotes listing 5-digit Hawaii/Alaska ZIPs that take extended zones.

Zone codes encode service + zone: Ground 002-008 (zone 2-8), 3 Day Select
302-308, 2nd Day Air 202-208, 2nd Day Air A.M. 242-248, Next Day Air Saver
132-138, Next Day Air 102-108. Codes outside those runs (045, 125, 225 for
Puerto Rico; 44/46, 124/126, 224/226 for AK/HI footnotes) are "extended"
zones. "-" means the service is not offered to that destination.
"""

import io
import re

SERVICES = [
    {"id": "ground",     "name": "UPS Ground",            "short": "Ground"},
    {"id": "three_day",  "name": "UPS 3 Day Select",      "short": "3 Day Select"},
    {"id": "two_day",    "name": "UPS 2nd Day Air",       "short": "2nd Day Air"},
    {"id": "two_day_am", "name": "UPS 2nd Day Air A.M.",  "short": "2nd Day A.M."},
    {"id": "nda_saver",  "name": "UPS Next Day Air Saver", "short": "NDA Saver"},
    {"id": "nda",        "name": "UPS Next Day Air",      "short": "Next Day Air"},
]

# Typical business days in transit, by service and (for Ground) zone.
GROUND_DAYS = {2: "1 day", 3: "2 days", 4: "2–3 days", 5: "3 days",
               6: "3–4 days", 7: "4 days", 8: "4–5 days"}
FLAT_DAYS = {"three_day": "3 days", "two_day": "2 days", "two_day_am": "2 days (a.m.)",
             "nda_saver": "next day (eve)", "nda": "next day"}

_HEADER_ALIASES = {
    "ground": "ground",
    "3 day select": "three_day",
    "2nd day air": "two_day",
    "2nd day air a.m.": "two_day_am",
    "next day air saver": "nda_saver",
    "next day air": "nda",
}


def _load_rows(path=None, blob=None):
    """Return all sheet rows as tuples. Sniffs the real format by magic bytes
    (UPS serves xlsx content with a .xls name)."""
    if blob is None:
        with open(path, "rb") as f:
            blob = f.read()
    if blob[:4] == b"PK\x03\x04":  # OOXML zip container -> xlsx
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        return [tuple(c for c in row) for row in ws.iter_rows(values_only=True)]
    if blob[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":  # legacy BIFF .xls
        import xlrd
        wb = xlrd.open_workbook(file_contents=blob)
        sh = wb.sheet_by_index(0)
        return [tuple(sh.cell_value(r, c) or None for c in range(sh.ncols))
                for r in range(sh.nrows)]
    raise ValueError("Not a recognizable Excel file (.xls or .xlsx)")


def _tier(code):
    """Map a raw zone code to a display tier: 2-8, 'ext', or None."""
    if code is None or code == "-":
        return None
    n = int(code)
    last = n % 10
    base = n - last
    if last in range(2, 9) and base in (0, 100, 130, 200, 240, 300):
        return last
    return "ext"


def parse_chart(path=None, blob=None):
    rows = _load_rows(path=path, blob=blob)

    origin = None
    header_idx = None
    col_map = None  # column index -> service id
    zones = {}
    notes = []

    for i, row in enumerate(rows):
        c0 = row[0]
        text = str(c0).strip() if c0 is not None else ""

        if origin is None:
            m = re.search(r"originating in ZIP Codes?\s+(\d{3})", text)
            if m:
                origin = m.group(1)

        if col_map is None and text.lower().startswith("dest"):
            header_idx = i
            col_map = {}
            for j, cell in enumerate(row[1:], start=1):
                if cell is None:
                    continue
                key = str(cell).strip().lower()
                if key in _HEADER_ALIASES:
                    col_map[j] = _HEADER_ALIASES[key]
            continue

        if col_map and re.fullmatch(r"\d{3}", text):
            entry = {}
            for j, svc in col_map.items():
                raw = row[j]
                raw = str(raw).strip() if raw is not None else "-"
                entry[svc] = {"code": raw if raw != "-" else None, "tier": _tier(raw)}
            zones[text] = entry

    if not zones or origin is None:
        raise ValueError("Workbook does not look like a UPS zone chart")

    exceptions = _parse_footnotes(rows)

    # AK/HI prefixes are absent from the main table; derive prefix-level rows
    # from the footnote ZIPs so the map can shade them as extended.
    for z5 in exceptions:
        prefix = z5[:3]
        if prefix not in zones:
            zones[prefix] = {
                svc["id"]: {"code": None,
                            "tier": "ext" if svc["id"] in ("ground", "two_day", "nda") else None}
                for svc in SERVICES
            }

    return {
        "origin": {"prefix": origin},
        "services": [
            {**svc,
             "days": GROUND_DAYS if svc["id"] == "ground" else None,
             "flat_days": FLAT_DAYS.get(svc["id"])}
            for svc in SERVICES
        ],
        "zones": zones,
        "exceptions": exceptions,
        "notes": notes,
    }


def _parse_footnotes(rows):
    """Collect 5-digit AK/HI exception ZIPs -> {'ground': 44, 'nda': 124, 'two_day': 224}."""
    exceptions = {}
    current = None
    for row in rows:
        cells = [c for c in row if c is not None]
        if not cells:
            continue
        first = str(cells[0]).strip()
        m = re.search(r"Zone\s+(\d+)\s+for Ground,?\s+Zone\s+(\d+)\s+for Next Day Air"
                      r"\s+and Zone\s+(\d+)\s+for 2nd Day Air", first)
        if m:
            current = {"ground": int(m.group(1)), "nda": int(m.group(2)),
                       "two_day": int(m.group(3))}
            continue
        if current is not None:
            zips = [str(int(c)) for c in cells
                    if isinstance(c, (int, float)) or str(c).strip().isdigit()]
            if zips:
                for z in zips:
                    exceptions[z] = current
            elif not re.match(r"^\d", first):
                # a non-numeric, non-header row ends the current block unless
                # it introduces another zone sentence (handled above)
                if "Zone" not in first:
                    current = None
    return exceptions
